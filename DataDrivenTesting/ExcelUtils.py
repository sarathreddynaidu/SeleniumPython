import openpyxl


def getRowCount(filePath, sheetName):
    workbook = openpyxl.load_workbook(filePath)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.max_row


def getColumnCount(filePath, sheetName):
    workbook = openpyxl.load_workbook(filePath)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.max_column


def readData(filePath, sheetName, rowNum, columnNum):
    workbook = openpyxl.load_workbook(filePath)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rowNum, column=columnNum).value


def writeData(filePath, sheetName, rowNum, columnNum, data):
    workbook = openpyxl.load_workbook(filePath)
    sheet = workbook.get_sheet_by_name(sheetName)
    sheet.cell(row=rowNum, column=columnNum).value = data
    workbook.save(filePath)
