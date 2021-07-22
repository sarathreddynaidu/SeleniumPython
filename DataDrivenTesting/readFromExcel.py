import openpyxl

filePath = "C:\SeleniumPython\DataDrivenTesting\SampleData.xlsx"

workbook = openpyxl.load_workbook(filePath)
sheet = workbook.active
# If there are multiple sheets use - workbook.get_sheet_by_name("Sheet1")

rows = sheet.max_row
columns = sheet.max_column

print("Number of rows: ", rows)
print("Number of columns: ", columns)

for r in range(1, rows+1):
    for c in range(1, columns+1):
        print(sheet.cell(row=r, column=c).value, end="   ")
    print()
